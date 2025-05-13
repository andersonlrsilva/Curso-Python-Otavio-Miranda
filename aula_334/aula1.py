
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active  # type: ignore

# Criando cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')
worksheet.cell(1, 4, 'Matriculado')


students = [
    # nome   idade nota
    ['Anderson',   14, 5.5, 'yes'],
    ['João',       14, 5.5, 'Pré-matricula'],
    ['Maria',      13, 9.7,  'Aguardando'],
    ['Luiz',       15, 8.8, 'Sim'],
    ['Alberto',    14, 5.5, 'Não'],

]

# Preenchendo linhs e colunas modo 1
# for i, studend_row in enumerate(students, start=2):
#     for j, student_collumn in enumerate(studend_row, start=1):
#         worksheet.cell(i, j, student_collumn)


# Preenchendo linhas e colunas modo 2
for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)
